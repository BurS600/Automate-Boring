import openpyxl,os

os.chdir(r"C:\Users\burha\Documents\Curriculum Vitae\CareerRelated\JobSearch2022")

print(os.getcwd())


workbook = openpyxl.load_workbook('Company_Roles_List.xlsx')

# similar to open(), use load_workbook() ; workbook object returned

print(workbook.get_sheet_names())

# to see all the sheet names

sheet = workbook['Sheet1']

# returns sheet object

print(sheet['A1'].value)



# cell = sheet['A1']

# print(cell.value)

# returns cell object ; can just use cell.value method in one go without needing to store in a cell object
# just cast datatype if needed as reading excel the datatype specified originally in excel



# you can use the following to specify cells using row and column numbers beginning at 1


print(sheet.cell(row=1, column=1).value)

for i in range(4,12):

    print(sheet.cell(row=i,column=3).value)

# iterating over a column and returning desired cells






