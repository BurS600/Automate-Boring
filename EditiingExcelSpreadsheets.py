import openpyxl, os

wb = openpyxl.Workbook()

# in this example, instead of loading an existing workbook
# we are initialising a new workbook using .Workbook()

print(wb.get_sheet_names())

sheet = wb['Sheet']

print(sheet)

# returns sheet object and stores it in variable

print(sheet['A1'].value == None)

sheet['A1'] = 42
sheet['A2'] = "Hello"


os.chdir(r"C:\Users\burha\Documents\Curriculum Vitae\CareerRelated\JobSearch2022")
print(os.getcwd())
         
wb.save('openpyxleditingexample.xlsx')

# when using load_workbook() and making changes, use .save() and
# think about saving with a new file name so you can retain original file

sheet2 = wb.create_sheet()
sheet2.title = 'My New Sheet Name'

# creating a new sheet and changing it's title using .title

wb.save('example2.xlsx')

print(wb.get_sheet_names())


wb.create_sheet(index=0, title='mySheetwithindex0')

# using keyword arguments to moving position of sheet and naming it

wb.save('example2.xlsx')
